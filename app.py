from flask import Flask, request, jsonify, render_template, url_for, send_file
import os
import cv2
import numpy as np
import tensorflow as tf
from openpyxl import load_workbook, Workbook
from PIL import Image
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Load the trained model
model = tf.keras.models.load_model('mnist_model.h5')

# Global variables
image_files = []
current_image_index = 0
current_folder = ""

# Set up upload folder
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure the upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Allowed file extensions
ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp'}

# Coordinates to extract digits from the image
coordinates = [
    (325, 227, 359, 260),  # Q.No a
    (365, 228, 395, 259),  # Q.No b
    (401, 228, 429, 259),  # Q.No c
    (436, 227, 470, 259),  # Q.No d
    (473, 229, 506, 258),  # Q.No e
    (508, 228, 541, 259),  # Q.No f
    (544, 228, 577, 259),  # Q.No g
    (581, 228, 615, 259),  # Q.No h
    (619, 230, 654, 259),  # Q.No i
    (657, 229, 688, 259),  # Q.No j
    # New coordinates for 2a, 2b, 3a, 3b, etc.
    (77, 315, 111, 345), (118, 319, 147, 345),  # 2a, 2b
    (325, 318, 358, 344), (362, 317, 396, 344),  # 3a, 3b
    (77, 351, 110, 376), (119, 351, 148, 379),  # 4a, 4b
    (329, 350, 358, 378), (363, 350, 396, 378),  # 5a, 5b
    (78, 383, 113, 413), (117, 383, 148, 413),  # 6a, 6b
    (325, 383, 359, 411), (362, 383, 398, 410),  # 7a, 7b
    (78, 416, 114, 444), (117, 416, 149, 444),  # 8a, 8b
    (325, 415, 359, 444), (362, 415, 398, 444)  # 9a, 9b
    # Add your existing coordinates here
]

# Helper functions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_blank_region(region, intensity_threshold=200, variance_threshold=50):
    """
    Check if the region is blank based on mean intensity and variance.

    Parameters:
    - region: The image region to check.
    - intensity_threshold: The mean intensity threshold to consider the region as blank.
    - variance_threshold: The variance threshold to consider the region as blank.

    Returns:
    - bool: True if the region is blank, False otherwise.
    """
    mean_intensity = np.mean(region)
    variance = np.var(region)

    return mean_intensity > intensity_threshold and variance < variance_threshold

def preprocess_image(image):
    img = Image.fromarray(image).convert('L')  # Convert to grayscale
    img = img.resize((28, 28))  # Resize to 28x28
    img_array = np.array(img).astype('float32') / 255  # Normalize
    img_array = 1 - img_array  # Invert colors if necessary
    img_array = img_array.reshape(1, 28, 28, 1)  # Reshape to match model input
    return img_array

def predict_digits(image_path, coordinates):
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        raise FileNotFoundError(f"Image at path {image_path} not found or could not be opened.")
    
    predictions = []
    for (x1, y1, x2, y2) in coordinates:
        digit_img = img[y1:y2, x1:x2]
        if digit_img.size == 0:
            predictions.append("")
            continue
        
        # Check if the region is mostly white and has low variance (indicating it's blank)
        if is_blank_region(digit_img):
            predictions.append("")
        else:
            # Preprocess the digit image
            input_image = preprocess_image(digit_img)
            prediction = model.predict(input_image)
            predicted_digit = np.argmax(prediction, axis=1)[0]
            predictions.append(str(predicted_digit))
    
    return predictions
    # Add your prediction logic here
    

def write_to_excel(predictions, row_num, headings, excel_path):
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(headings)
    else:
        wb = load_workbook(excel_path)
        ws = wb.active

    ws.append([row_num] + predictions)
    wb.save(excel_path)
    # Add your logic to write predictions to Excel here
    

@app.route('/')
def index():
    return render_template('index.html')
def allowed_file(filename):
    allowed_extensions = {'bmp', 'jpg', 'jpeg', 'png', 'gif'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

@app.route('/choose-folder', methods=['POST'])
def choose_folder():
    global image_files, current_image_index, current_folder

    if 'folder' not in request.files:
        return jsonify({'status': 'error', 'message': 'No folder part'})
    
    files = request.files.getlist('folder')
    
    # Debug: Print the number of files received
    print(f"Number of files received: {len(files)}")

    if not files or files[0].filename == '':
        return jsonify({'status': 'error', 'message': 'No selected folder'})

    folder_name = secure_filename(files[0].filename.split('/')[0])
    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    
    # Debug: Print the folder path
    print(f"Saving files to: {folder_path}")

    os.makedirs(folder_path, exist_ok=True)

    saved_files = []
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(folder_path, filename)
            
            # Debug: Print file path before saving
            print(f"Saving file: {file_path}")
            
            file.save(file_path)
            saved_files.append(filename)

    # Update global variables
    current_folder = folder_path
    image_files = [f for f in saved_files if allowed_file(f)]
    current_image_index = 0

    if image_files:
        image_path = url_for('static', filename=f'uploads/{folder_name}/{image_files[0]}')
        return jsonify({'status': 'success', 'image_path': image_path})
    else:
        return jsonify({'status': 'error', 'message': 'No images found in the uploaded folder'})




@app.route('/next-image', methods=['GET'])
def next_image():
    global current_image_index
    if current_image_index < len(image_files) - 1:
        current_image_index += 1
        image_path = url_for('static', filename=f'uploads/{os.path.basename(current_folder)}/{image_files[current_image_index]}')
        return jsonify({'image_path': image_path})
    else:
        return jsonify({'image_path': None})

@app.route('/prev-image', methods=['GET'])
def prev_image():
    global current_image_index
    if current_image_index > 0:
        current_image_index -= 1
        image_path = url_for('static', filename=f'uploads/{os.path.basename(current_folder)}/{image_files[current_image_index]}')
        return jsonify({'image_path': image_path})
    else:
        return jsonify({'image_path': None})

@app.route('/submit-image', methods=['POST'])
def submit_image():
    data = request.json
    image_path = data.get('image_path')
    
    if image_path.startswith(request.url_root):
        image_path = image_path[len(request.url_root):]
    
    full_image_path = os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(current_folder), os.path.basename(image_path))
    try:
        predictions = predict_digits(full_image_path, coordinates)
    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    
    row_num = current_image_index + 1
    headings = ['S.No', '1a', '1b', '1c', '1d', '1e', '1f', '1g', '1h', '1i', '1j', 
                '2a', '2b', '3a', '3b', '4a', '4b', '5a', '5b', '6a', '6b', '7a', 
                '7b', '8a', '8b', '9a', '9b']
    excel_path = os.path.join(current_folder, 'digit_predictions.xlsx')
    write_to_excel(predictions, row_num, headings, excel_path)
    
    return jsonify({'status': 'success', 'predictions': predictions})

@app.route('/download-excel', methods=['GET'])
def download_excel():
    excel_path = os.path.join(current_folder, 'digit_predictions.xlsx')
    if os.path.exists(excel_path):
        return send_file(excel_path, as_attachment=True)
    else:
        return jsonify({'status': 'error', 'message': 'Excel file not found.'}), 404

if __name__ == '__main__':
    app.run(debug=True) 